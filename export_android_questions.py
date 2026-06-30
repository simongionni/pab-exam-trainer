import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "pab-s1-quiz.xlsx"
OUTPUT_PATH = ROOT / "android_app" / "app" / "src" / "main" / "assets" / "questions.json"


def normalize_answer(value: str) -> str:
    return "".join(sorted(set(re.sub(r"[^A-F]", "", str(value or "").upper()))))


def split_question_text(text: str) -> tuple[str, list[dict[str, str]]]:
    lines = [line.rstrip() for line in str(text or "").splitlines()]
    prompt_lines: list[str] = []
    options: list[tuple[str, list[str]]] = []
    current = None

    option_re = re.compile(r"^\s*([A-F])[\.\)]\s*(.*)$")
    for line in lines:
        match = option_re.match(line)
        if match:
            current = [match.group(1), [match.group(2).strip()]]
            options.append(current)
            continue
        if current is not None:
            current[1].append(line.strip())
        else:
            prompt_lines.append(line.strip())

    prompt = "\n".join(line for line in prompt_lines if line).strip()
    clean_options = []
    for letter, parts in options:
        clean_options.append(
            {
                "letter": letter,
                "text": " ".join(part for part in parts if part).strip(),
            }
        )
    return prompt, clean_options


def export_questions() -> int:
    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    sheet = workbook["Quiz"]
    questions = []

    for row in range(2, sheet.max_row + 1):
        qid = str(sheet.cell(row, 1).value or "").strip()
        source = str(sheet.cell(row, 2).value or "").strip()
        original = str(sheet.cell(row, 3).value or "").strip()
        text = sheet.cell(row, 4).value
        correct = normalize_answer(sheet.cell(row, 5).value)
        prompt, options = split_question_text(text)

        if not qid or not prompt or not options or not correct:
            continue

        questions.append(
            {
                "id": qid,
                "source": source,
                "originalNumber": original,
                "prompt": prompt,
                "options": options,
                "correctLetters": list(correct),
            }
        )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(
        json.dumps(questions, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return len(questions)


if __name__ == "__main__":
    count = export_questions()
    print(f"Esportate {count} domande in {OUTPUT_PATH}")
