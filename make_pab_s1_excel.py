import pathlib
import re
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


answers = {
    1: "D", 2: "C", 3: "D", 4: "D", 5: "C", 6: "C", 7: "C", 8: "A", 9: "A",
    10: "ADE", 11: "BCE", 12: "AC", 13: "D", 14: "A", 15: "A", 16: "D",
    17: "C", 18: "B", 19: "D", 20: "D", 21: "A", 22: "B", 23: "B",
    24: "B", 25: "A", 26: "AB", 27: "A", 28: "A", 29: "A", 30: "AB",
    31: "B", 32: "C", 33: "B", 34: "D", 35: "C", 36: "A", 37: "BCE",
    38: "B", 39: "CD", 40: "C", 41: "AD", 42: "B", 43: "C", 44: "AD",
    45: "A", 46: "A", 47: "A", 48: "AD", 49: "B", 50: "B", 51: "B",
    52: "D", 53: "AB", 54: "A", 55: "B", 56: "C", 57: "C", 58: "B",
    59: "B", 60: "A", 61: "C", 62: "D", 63: "D", 64: "B", 65: "A",
}

notes = {
    49: "Formula OCR poco leggibile nel PDF; lettera corretta presa dall'evidenziazione.",
    64: "Formula OCR poco leggibile nel PDF; lettera corretta presa dall'evidenziazione.",
}


def check_formula(row: int) -> str:
    cleaned = f'SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(D{row},",",""),";","")," ","")'
    return f'=IF(D{row}="","",IF(UPPER({cleaned})=C{row},"OK","NO"))'


def parse_questions(path: pathlib.Path) -> list[dict]:
    lines = path.read_text(encoding="utf-8").splitlines()
    questions: list[dict] = []
    current = None

    for raw in lines:
        line = raw.strip()
        if not line or line.startswith("=====") or line in {"Salesforce Certified Platform", "App Builder"}:
            continue

        line_for_number = re.sub(r"^IO\b", "10", line)
        match = re.match(r"^(\d{1,2})(?:\s+(.+))?$", line_for_number)
        if match and 1 <= int(match.group(1)) <= 65:
            number = int(match.group(1))
            rest = (match.group(2) or "").strip()

            if (
                number == 5
                and questions
                and questions[-1]["lines"]
                and questions[-1]["lines"][-1].startswith(
                    "Sales reps at Universal Containers create multiple quotes"
                )
            ):
                preface = questions[-1]["lines"].pop()
                current = {"num": number, "lines": [preface]}
            else:
                current = {"num": number, "lines": []}

            if rest:
                current["lines"].append(rest)
            questions.append(current)
            continue

        if current is not None:
            current["lines"].append(line)

    for question in questions:
        text = "\n".join(question["lines"])
        text = re.sub(r"\n([A-E])\.\n", r"\n\1. ", text)
        text = re.sub(r"[ \t]+", " ", text)
        question["text"] = text.strip()

    return questions


def merged_ocr_lines(json_path: pathlib.Path) -> list[str]:
    pages = json.loads(json_path.read_text(encoding="utf-8-sig"))
    output: list[str] = []

    for page in pages:
        output.append(f"===== PAGE {page['page']:02} =====")
        raw_lines = []
        for line in page["lines"]:
            if not line["words"]:
                continue
            x = min(word["x"] for word in line["words"])
            y = min(word["y"] for word in line["words"])
            raw_lines.append({"text": line["text"].strip(), "x": x, "y": y})

        groups: list[list[dict]] = []
        for line in sorted(raw_lines, key=lambda item: (item["y"], item["x"])):
            if groups and abs(line["y"] - groups[-1][0]["y"]) <= 8:
                groups[-1].append(line)
            else:
                groups.append([line])

        for group in groups:
            parts = [item["text"] for item in sorted(group, key=lambda item: item["x"]) if item["text"]]
            output.append(" ".join(parts))
        output.append("")

    return output


def main() -> None:
    merged_path = pathlib.Path("pab-s1-ocr-lines-fixed.txt")
    json_path = pathlib.Path("pab-s1-ocr-words.json")
    if json_path.exists():
        merged_path.write_text("\n".join(merged_ocr_lines(json_path)), encoding="utf-8")

    questions = parse_questions(merged_path if merged_path.exists() else pathlib.Path("pab-s1-ocr-lines.txt"))
    by_number = {question["num"]: question for question in questions}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quiz"
    sheet.append(["N", "Domanda e opzioni (OCR)", "Corretta", "La tua risposta", "Esito", "Note"])

    for number in range(1, 66):
        question = by_number.get(number)
        text = question["text"] if question else "[Da controllare: parsing OCR non riuscito]"
        note = notes.get(number, "")
        if question is None:
            note = "Parsing OCR da verificare"
        row = sheet.max_row + 1
        sheet.append([number, text, answers.get(number, ""), "", check_formula(row), note])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = {"A": 6, "B": 100, "C": 12, "D": 16, "E": 10, "F": 35}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[2].alignment = Alignment(horizontal="center", vertical="top")
        row[3].alignment = Alignment(horizontal="center", vertical="top")
        row[4].alignment = Alignment(horizontal="center", vertical="top")
        row[5].alignment = Alignment(wrap_text=True, vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:F{sheet.max_row}"

    instructions = workbook.create_sheet("Istruzioni")
    instructions["A1"] = "Come usarlo"
    instructions["A1"].font = Font(bold=True, size=14)
    instructions["A2"] = (
        "Inserisci la tua risposta nella colonna D del foglio Quiz. "
        "Per domande a risposta multipla usa lettere in ordine alfabetico, es. AD o BCE. "
        "La colonna E mostra OK/NO."
    )
    instructions["A3"] = (
        "Le risposte corrette sono quelle evidenziate in giallo nel PDF originale. "
        "La trascrizione deriva da OCR, quindi alcune formule possono contenere errori visivi."
    )
    instructions.column_dimensions["A"].width = 120

    workbook.save("pab-s1-quiz.xlsx")
    print(f"saved pab-s1-quiz.xlsx with {len(questions)} parsed questions")


if __name__ == "__main__":
    main()
