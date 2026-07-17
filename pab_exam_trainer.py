import json
import random
import re
import tkinter as tk
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from tkinter import messagebox, ttk

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "pab-s1-quiz.xlsx"
STATS_PATH = ROOT / "pab_exam_stats.json"
SESSION_PATH = ROOT / "pab_exam_session.json"
SESSION_SIZE = 65
SUPABASE_URL = "https://haorkjzxxhzpcvmvaklk.supabase.co"
SUPABASE_ANON_KEY = "sb_publishable_ljBw3GVH83nOhfVB-uXF2g_vek6Ev4C"


@dataclass
class Question:
    qid: str
    source: str
    original_num: str
    prompt: str
    options: list[tuple[str, str]]
    correct_letters: set[str]


class SupabaseError(RuntimeError):
    pass


class SupabaseClient:
    def __init__(self, url: str, anon_key: str) -> None:
        self.url = url.rstrip("/")
        self.anon_key = anon_key
        self.access_token = ""
        self.refresh_token = ""
        self.user_id = ""
        self.email = ""

    @property
    def online(self) -> bool:
        return bool(self.access_token and self.user_id)

    def load_saved_session(self) -> bool:
        if not SESSION_PATH.exists():
            return False
        try:
            data = json.loads(SESSION_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return False
        self.access_token = data.get("access_token", "")
        self.refresh_token = data.get("refresh_token", "")
        self.user_id = data.get("user_id", "")
        self.email = data.get("email", "")
        return bool(self.access_token and self.refresh_token)

    def save_session(self) -> None:
        SESSION_PATH.write_text(
            json.dumps(
                {
                    "access_token": self.access_token,
                    "refresh_token": self.refresh_token,
                    "user_id": self.user_id,
                    "email": self.email,
                },
                indent=2,
            ),
            encoding="utf-8",
        )

    def clear_session(self) -> None:
        self.access_token = ""
        self.refresh_token = ""
        self.user_id = ""
        self.email = ""
        if SESSION_PATH.exists():
            SESSION_PATH.unlink()

    def sign_in(self, email: str, password: str) -> None:
        data = self._request(
            "POST",
            "/auth/v1/token?grant_type=password",
            payload={"email": email, "password": password},
            auth=False,
        )
        self._apply_auth_response(data)

    def refresh_session(self) -> None:
        if not self.refresh_token:
            raise SupabaseError("Sessione Supabase mancante.")
        data = self._request(
            "POST",
            "/auth/v1/token?grant_type=refresh_token",
            payload={"refresh_token": self.refresh_token},
            auth=False,
        )
        self._apply_auth_response(data)

    def fetch_questions(self) -> list[Question]:
        rows = self._request(
            "GET",
            "/rest/v1/questions?"
            + urllib.parse.urlencode(
                {
                    "select": "id,source,original_number,prompt,options,correct_letters",
                    "order": "id.asc",
                }
            ),
        )
        questions = []
        for row in rows:
            options = [
                (str(option.get("letter", "")), str(option.get("text", "")))
                for option in row.get("options", [])
            ]
            correct = {str(letter) for letter in row.get("correct_letters", [])}
            if row.get("prompt") and options and correct:
                questions.append(
                    Question(
                        qid=str(row["id"]),
                        source=str(row.get("source") or ""),
                        original_num=str(row.get("original_number") or ""),
                        prompt=str(row["prompt"]),
                        options=options,
                        correct_letters=correct,
                    )
                )
        return questions

    def fetch_stats(self) -> dict:
        rows = self._request(
            "GET",
            "/rest/v1/question_stats?"
            + urllib.parse.urlencode(
                {
                    "select": "question_id,seen,correct,wrong",
                    "user_id": f"eq.{self.user_id}",
                }
            ),
        )
        return {
            "questions": {
                str(row["question_id"]): {
                    "seen": int(row.get("seen") or 0),
                    "correct": int(row.get("correct") or 0),
                    "wrong": int(row.get("wrong") or 0),
                }
                for row in rows
            }
        }

    def upsert_stat(self, qid: str, values: dict) -> None:
        self.upsert_stats({qid: values})

    def upsert_stats(self, question_values: dict) -> None:
        if not question_values:
            return
        payload = [
            {
                "user_id": self.user_id,
                "question_id": qid,
                "seen": int(values.get("seen", 0)),
                "correct": int(values.get("correct", 0)),
                "wrong": int(values.get("wrong", 0)),
            }
            for qid, values in question_values.items()
        ]
        self._request(
            "POST",
            "/rest/v1/question_stats?on_conflict=user_id,question_id",
            payload=payload,
            prefer="resolution=merge-duplicates,return=minimal",
            expect_json=False,
        )

    def update_question(self, question: Question) -> None:
        payload = {
            "prompt": question.prompt,
            "options": [
                {"letter": letter, "text": text}
                for letter, text in question.options
            ],
            "correct_letters": sorted(question.correct_letters),
        }
        self._request(
            "PATCH",
            "/rest/v1/questions?" + urllib.parse.urlencode({"id": f"eq.{question.qid}"}),
            payload=payload,
            prefer="return=minimal",
            expect_json=False,
        )

    def _apply_auth_response(self, data: dict) -> None:
        user = data.get("user") or {}
        self.access_token = data.get("access_token", "")
        self.refresh_token = data.get("refresh_token", self.refresh_token)
        self.user_id = user.get("id", self.user_id)
        self.email = user.get("email", self.email)
        if not self.access_token or not self.user_id:
            raise SupabaseError("Risposta login Supabase incompleta.")
        self.save_session()

    def _request(
        self,
        method: str,
        path: str,
        payload: object | None = None,
        auth: bool = True,
        prefer: str | None = None,
        expect_json: bool = True,
    ):
        body = None
        headers = {
            "apikey": self.anon_key,
            "Content-Type": "application/json",
        }
        if auth and self.access_token:
            headers["Authorization"] = f"Bearer {self.access_token}"
        if prefer:
            headers["Prefer"] = prefer
        if payload is not None:
            body = json.dumps(payload).encode("utf-8")

        request = urllib.request.Request(
            self.url + path,
            data=body,
            headers=headers,
            method=method,
        )
        try:
            with urllib.request.urlopen(request, timeout=20) as response:
                raw = response.read().decode("utf-8")
        except urllib.error.HTTPError as error:
            detail = error.read().decode("utf-8", errors="replace")
            raise SupabaseError(f"Supabase HTTP {error.code}: {detail}") from error
        except urllib.error.URLError as error:
            raise SupabaseError(f"Supabase non raggiungibile: {error.reason}") from error

        if not expect_json or not raw.strip():
            return None
        return json.loads(raw)


def normalize_answer(value: str) -> str:
    return "".join(sorted(set(re.sub(r"[^A-F]", "", str(value or "").upper()))))


def split_question_text(text: str) -> tuple[str, list[tuple[str, str]]]:
    lines = [line.rstrip() for line in str(text or "").splitlines()]
    prompt_lines: list[str] = []
    options: list[tuple[str, list[str]]] = []
    current = None

    option_re = re.compile(r"^\s*([A-F])[\.\)]\s*(.*)$")
    for line in lines:
        match = option_re.match(line)
        if match:
            current = [match.group(1), [match.group(2).strip()]]
            options.append(current)
            continue
        if current is not None:
            current[1].append(line.strip())
        else:
            prompt_lines.append(line.strip())

    prompt = "\n".join(line for line in prompt_lines if line).strip()
    clean_options = []
    for letter, parts in options:
        value = " ".join(part for part in parts if part).strip()
        clean_options.append((letter, value))
    return prompt, clean_options


def load_questions() -> list[Question]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    sheet = workbook["Quiz"]
    questions: list[Question] = []

    for row in range(2, sheet.max_row + 1):
        qid = str(sheet.cell(row, 1).value)
        source = str(sheet.cell(row, 2).value)
        original = str(sheet.cell(row, 3).value)
        text = sheet.cell(row, 4).value
        correct = normalize_answer(sheet.cell(row, 5).value)
        prompt, options = split_question_text(text)
        if not prompt or not options or not correct:
            continue
        questions.append(
            Question(
                qid=qid,
                source=source,
                original_num=original,
                prompt=prompt,
                options=options,
                correct_letters=set(correct),
            )
        )
    return questions


def load_questions_from_workbook() -> list[Question]:
    return load_questions()


def load_stats() -> dict:
    if not STATS_PATH.exists():
        return {"questions": {}}
    try:
        return json.loads(STATS_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"questions": {}}


def save_stats(stats: dict) -> None:
    STATS_PATH.write_text(json.dumps(stats, indent=2, ensure_ascii=False), encoding="utf-8")


def question_stats(stats: dict, qid: str) -> dict:
    return stats.setdefault("questions", {}).setdefault(
        qid, {"seen": 0, "correct": 0, "wrong": 0}
    )


def merge_stats(local_stats: dict, remote_stats: dict) -> dict:
    merged = {"questions": {}}
    all_qids = set(local_stats.get("questions", {})) | set(remote_stats.get("questions", {}))
    for qid in all_qids:
        local = question_stats(local_stats, qid)
        remote = question_stats(remote_stats, qid)
        merged["questions"][qid] = {
            "seen": max(int(local.get("seen", 0)), int(remote.get("seen", 0))),
            "correct": max(int(local.get("correct", 0)), int(remote.get("correct", 0))),
            "wrong": max(int(local.get("wrong", 0)), int(remote.get("wrong", 0))),
        }
    return merged


def eligible_questions(candidates: list[Question], stats: dict) -> list[Question]:
    counts = [question_stats(stats, q.qid)["seen"] for q in candidates]
    min_count = min(counts)
    if min_count == 0:
        return [q for q in candidates if question_stats(stats, q.qid)["seen"] == 0]
    return [q for q in candidates if question_stats(stats, q.qid)["seen"] < 3 * min_count]


def build_session(questions: list[Question], stats: dict) -> list[Question]:
    remaining = questions[:]
    picked: list[Question] = []

    while remaining and len(picked) < min(SESSION_SIZE, len(questions)):
        eligible = eligible_questions(remaining, stats)
        pool = eligible if eligible else remaining
        choice = random.choice(pool)
        picked.append(choice)
        remaining.remove(choice)

    return picked


class ExamTrainer(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("PAB Exam Trainer")
        self.geometry("1000x760")
        self.minsize(820, 620)

        self.sync = SupabaseClient(SUPABASE_URL, SUPABASE_ANON_KEY)
        self.questions = load_questions_from_workbook()
        if len(self.questions) < SESSION_SIZE:
            messagebox.showwarning(
                "Domande insufficienti",
                f"Ho trovato solo {len(self.questions)} domande valide.",
            )
        self.stats = load_stats()
        self.session: list[Question] = []
        self.index = 0
        self.current_options: list[tuple[str, str, str]] = []
        self.answer_vars: dict[str, tk.BooleanVar] = {}
        self.option_widgets: dict[str, ttk.Checkbutton] = {}
        self.feedback_shown = False
        self.correct_in_session = 0

        self._build_ui()
        self.restore_saved_supabase_session()
        self.new_session()

    def _build_ui(self) -> None:
        top = ttk.Frame(self, padding=12)
        top.pack(fill="x")

        self.progress_label = ttk.Label(top, text="")
        self.progress_label.pack(side="left")

        ttk.Button(top, text="Nuovo esame", command=self.confirm_new_session).pack(side="right")
        ttk.Button(top, text="Modifica domanda", command=self.edit_current_question).pack(side="right", padx=(0, 8))
        ttk.Button(top, text="Statistiche", command=self.show_stats).pack(side="right", padx=(0, 8))
        ttk.Button(top, text="Reset counter", command=self.reset_counters).pack(side="right", padx=(0, 8))
        ttk.Button(top, text="Sync", command=self.sync_now).pack(side="right", padx=(0, 8))
        ttk.Button(top, text="Logout", command=self.logout).pack(side="right", padx=(0, 8))
        ttk.Button(top, text="Login", command=self.login).pack(side="right", padx=(0, 8))

        sync_row = ttk.Frame(self, padding=(12, 0, 12, 8))
        sync_row.pack(fill="x")
        self.sync_status_label = ttk.Label(sync_row, text="Sync: offline")
        self.sync_status_label.pack(side="left")

        body = ttk.Frame(self, padding=(12, 0, 12, 12))
        body.pack(fill="both", expand=True)

        self.prompt = tk.Text(body, wrap="word", height=11, font=("Segoe UI", 11))
        self.prompt.configure(state="disabled")
        self.prompt.pack(fill="x", pady=(0, 10))

        self.options_frame = ttk.Frame(body)
        self.options_frame.pack(fill="both", expand=True)

        self.feedback = tk.Text(body, wrap="word", height=8, font=("Segoe UI", 10))
        self.feedback.configure(state="disabled")
        self.feedback.pack(fill="x", pady=(10, 10))

        bottom = ttk.Frame(body)
        bottom.pack(fill="x")

        self.submit_button = ttk.Button(bottom, text="Conferma risposta", command=self.submit_answer)
        self.submit_button.pack(side="left")
        self.next_button = ttk.Button(bottom, text="Prossima domanda", command=self.next_question)
        self.next_button.pack(side="left", padx=(8, 0))

    def restore_saved_supabase_session(self) -> None:
        if not self.sync.load_saved_session():
            self.update_sync_status("offline")
            return
        try:
            self.sync.refresh_session()
            self.sync_now(show_success=False)
        except SupabaseError:
            self.sync.clear_session()
            self.update_sync_status("offline")

    def login(self) -> None:
        credentials = self.ask_credentials()
        if credentials is None:
            return
        email, password = credentials
        try:
            self.sync.sign_in(email, password)
            self.sync_now(show_success=False)
        except SupabaseError as error:
            messagebox.showerror("Login Supabase", str(error))
            self.update_sync_status("offline")
            return
        messagebox.showinfo("Login Supabase", f"Login effettuato: {self.sync.email}")
        self.new_session()

    def logout(self) -> None:
        if not self.sync.online:
            self.sync.clear_session()
            self.update_sync_status("offline")
            return
        if not messagebox.askyesno("Logout", "Vuoi scollegare questo trainer da Supabase?"):
            return
        self.sync.clear_session()
        self.update_sync_status("offline")

    def sync_now(self, show_success: bool = True) -> None:
        if not self.sync.online:
            messagebox.showwarning("Sync", "Fai login a Supabase prima di sincronizzare.")
            return
        try:
            remote_questions = self.sync.fetch_questions()
            remote_stats = self.sync.fetch_stats()
        except SupabaseError as error:
            self.update_sync_status("offline, ultimo sync fallito")
            if show_success:
                messagebox.showerror("Sync", str(error))
            return

        if remote_questions:
            self.questions = remote_questions
        self.stats = merge_stats(self.stats, remote_stats)
        save_stats(self.stats)
        try:
            self.sync.upsert_stats(self.stats.get("questions", {}))
        except SupabaseError as error:
            self.update_sync_status("online, upload stats fallito")
            if show_success:
                messagebox.showerror("Sync", str(error))
            return
        self.update_sync_status(f"online: {self.sync.email}")
        if show_success:
            messagebox.showinfo("Sync", "Statistiche sincronizzate con Supabase.")

    def sync_question_stat(self, qid: str) -> None:
        if not self.sync.online:
            return
        try:
            self.sync.upsert_stat(qid, question_stats(self.stats, qid))
            self.update_sync_status(f"online: {self.sync.email}")
        except SupabaseError:
            self.update_sync_status("offline, sync in sospeso")

    def update_sync_status(self, value: str) -> None:
        self.sync_status_label.configure(text=f"Sync: {value}")

    def ask_credentials(self) -> tuple[str, str] | None:
        dialog = tk.Toplevel(self)
        dialog.title("Login Supabase")
        dialog.transient(self)
        dialog.grab_set()
        dialog.resizable(False, False)

        frame = ttk.Frame(dialog, padding=16)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Email").grid(row=0, column=0, sticky="w")
        email_var = tk.StringVar(value=self.sync.email)
        email_entry = ttk.Entry(frame, textvariable=email_var, width=42)
        email_entry.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(4, 10))

        ttk.Label(frame, text="Password").grid(row=2, column=0, sticky="w")
        password_var = tk.StringVar()
        password_entry = ttk.Entry(frame, textvariable=password_var, show="*", width=42)
        password_entry.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(4, 14))

        result: dict[str, tuple[str, str] | None] = {"value": None}

        def submit() -> None:
            email = email_var.get().strip()
            password = password_var.get()
            if not email or not password:
                messagebox.showwarning("Login Supabase", "Inserisci email e password.", parent=dialog)
                return
            result["value"] = (email, password)
            dialog.destroy()

        ttk.Button(frame, text="Annulla", command=dialog.destroy).grid(row=4, column=0, sticky="e")
        ttk.Button(frame, text="Login", command=submit).grid(row=4, column=1, sticky="e", padx=(8, 0))

        dialog.bind("<Return>", lambda _event: submit())
        email_entry.focus_set()
        self.wait_window(dialog)
        return result["value"]

    def edit_current_question(self) -> None:
        if self.index >= len(self.session):
            messagebox.showwarning("Modifica domanda", "Non c'è una domanda attiva da modificare.")
            return
        if not self.sync.online:
            messagebox.showwarning("Modifica domanda", "Fai login a Supabase per salvare le modifiche nel database.")
            return

        question = self.session[self.index]
        dialog = tk.Toplevel(self)
        dialog.title(f"Modifica domanda {question.qid}")
        dialog.transient(self)
        dialog.grab_set()
        dialog.geometry("760x680")
        dialog.minsize(620, 520)

        frame = ttk.Frame(dialog, padding=16)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Testo della domanda").pack(anchor="w")
        prompt_editor = tk.Text(frame, wrap="word", height=8, font=("Segoe UI", 10))
        prompt_editor.insert("1.0", question.prompt)
        prompt_editor.pack(fill="x", pady=(4, 12))

        ttk.Label(frame, text="Risposte (seleziona quelle corrette)").pack(anchor="w")
        option_area = ttk.Frame(frame)
        option_area.pack(fill="both", expand=True, pady=(4, 10))
        option_entries: list[tuple[str, ttk.Entry]] = []
        correct_vars: dict[str, tk.BooleanVar] = {}
        for row, (letter, text) in enumerate(question.options):
            var = tk.BooleanVar(value=letter in question.correct_letters)
            correct_vars[letter] = var
            ttk.Checkbutton(option_area, text=letter, variable=var).grid(row=row, column=0, sticky="nw", padx=(0, 8), pady=4)
            entry = ttk.Entry(option_area)
            entry.insert(0, text)
            entry.grid(row=row, column=1, sticky="ew", pady=4)
            option_entries.append((letter, entry))
        option_area.columnconfigure(1, weight=1)

        count_var = tk.IntVar(value=len(question.correct_letters))
        count_row = ttk.Frame(frame)
        count_row.pack(fill="x", pady=(0, 12))
        ttk.Label(count_row, text="Numero di risposte corrette").pack(side="left")
        ttk.Spinbox(
            count_row,
            from_=1,
            to=len(question.options),
            width=5,
            textvariable=count_var,
        ).pack(side="left", padx=(8, 0))

        buttons = ttk.Frame(frame)
        buttons.pack(fill="x")

        def save() -> None:
            prompt = prompt_editor.get("1.0", "end").strip()
            options = [(letter, entry.get().strip()) for letter, entry in option_entries]
            correct = {letter for letter, var in correct_vars.items() if var.get()}
            try:
                expected_count = int(count_var.get())
            except (tk.TclError, ValueError):
                expected_count = 0
            if not prompt:
                messagebox.showwarning("Modifica domanda", "Il testo della domanda non può essere vuoto.", parent=dialog)
                return
            if any(not text for _, text in options):
                messagebox.showwarning("Modifica domanda", "Il testo di ogni risposta è obbligatorio.", parent=dialog)
                return
            if expected_count < 1 or expected_count > len(options):
                messagebox.showwarning("Modifica domanda", "Il numero di risposte corrette non è valido.", parent=dialog)
                return
            if len(correct) != expected_count:
                messagebox.showwarning(
                    "Modifica domanda",
                    f"Hai indicato {expected_count} risposte corrette, ma ne hai selezionate {len(correct)}.",
                    parent=dialog,
                )
                return

            updated = Question(question.qid, question.source, question.original_num, prompt, options, correct)
            try:
                self.sync.update_question(updated)
            except SupabaseError as error:
                messagebox.showerror("Modifica domanda", str(error), parent=dialog)
                return

            question.prompt = updated.prompt
            question.options = updated.options
            question.correct_letters = updated.correct_letters
            dialog.destroy()
            self.show_question(count_seen=False)
            messagebox.showinfo("Modifica domanda", "Domanda aggiornata nel database.")

        ttk.Button(buttons, text="Annulla", command=dialog.destroy).pack(side="right")
        ttk.Button(buttons, text="Salva nel database", command=save).pack(side="right", padx=(0, 8))
        prompt_editor.focus_set()

    def confirm_new_session(self) -> None:
        if messagebox.askyesno("Nuovo esame", "Vuoi iniziare una nuova sessione da 65 domande?"):
            self.new_session()

    def new_session(self) -> None:
        self.session = build_session(self.questions, self.stats)
        self.index = 0
        self.correct_in_session = 0
        self.show_question()

    def show_question(self, count_seen: bool = True) -> None:
        self.feedback_shown = False
        self.answer_vars.clear()
        self.option_widgets.clear()
        for child in self.options_frame.winfo_children():
            child.destroy()

        if self.index >= len(self.session):
            self.show_session_done()
            return

        question = self.session[self.index]
        if count_seen:
            question_stats(self.stats, question.qid)["seen"] += 1
            save_stats(self.stats)
            self.sync_question_stat(question.qid)

        self.current_options = self.shuffle_options(question)

        self.progress_label.configure(
            text=(
                f"Domanda {self.index + 1}/{len(self.session)}  "
                f"Score: {self.correct_in_session}/{self.index}  "
                f"Fonte: {question.source} #{question.original_num}"
            )
        )
        self.set_text(self.prompt, question.prompt)
        self.set_text(self.feedback, "")

        for display_letter, original_letter, text in self.current_options:
            var = tk.BooleanVar(value=False)
            self.answer_vars[display_letter] = var
            checkbox = ttk.Checkbutton(
                self.options_frame,
                text=f"{display_letter}. {text}",
                variable=var,
                command=self.enforce_answer_limit,
            )
            checkbox.pack(anchor="w", fill="x", pady=5)
            self.option_widgets[display_letter] = checkbox

        self.next_button.configure(state="disabled")
        self.enforce_answer_limit()

    def shuffle_options(self, question: Question) -> list[tuple[str, str, str]]:
        options = question.options[:]
        random.shuffle(options)
        display_letters = "ABCDEF"
        return [
            (display_letters[index], original_letter, text)
            for index, (original_letter, text) in enumerate(options)
        ]

    def selected_original_letters(self) -> set[str]:
        display_to_original = {display: original for display, original, _ in self.current_options}
        return {
            display_to_original[display]
            for display, var in self.answer_vars.items()
            if var.get()
        }

    def enforce_answer_limit(self) -> None:
        if self.index >= len(self.session):
            return
        limit = len(self.session[self.index].correct_letters)
        selected_count = sum(1 for var in self.answer_vars.values() if var.get())
        for display, widget in self.option_widgets.items():
            if self.answer_vars[display].get() or selected_count < limit:
                widget.configure(state="normal")
            else:
                widget.configure(state="disabled")
        self.submit_button.configure(state="normal" if selected_count == limit else "disabled")

    def submit_answer(self) -> None:
        if self.feedback_shown:
            return
        question = self.session[self.index]
        selected = self.selected_original_letters()
        correct = selected == question.correct_letters

        stats = question_stats(self.stats, question.qid)
        if correct:
            stats["correct"] += 1
            self.correct_in_session += 1
        else:
            stats["wrong"] += 1
        save_stats(self.stats)
        self.sync_question_stat(question.qid)

        selected_text = self.describe_letters(selected)
        correct_text = self.describe_letters(question.correct_letters)
        verdict = "CORRETTA" if correct else "SBAGLIATA"
        feedback = [
            verdict,
            "",
            f"La tua risposta: {selected_text or '(nessuna)'}",
            f"Risposta corretta: {correct_text}",
        ]
        self.set_text(self.feedback, "\n".join(feedback))
        self.feedback_shown = True
        self.submit_button.configure(state="disabled")
        self.next_button.configure(state="normal")

    def describe_letters(self, letters: set[str]) -> str:
        parts = []
        for display, original, text in self.current_options:
            if original in letters:
                parts.append(f"{display}. {text}")
        return "; ".join(parts)

    def next_question(self) -> None:
        self.index += 1
        self.show_question()

    def show_session_done(self) -> None:
        self.progress_label.configure(
            text=f"Sessione completata: {self.correct_in_session}/{len(self.session)} corrette"
        )
        self.set_text(self.prompt, "Sessione completata.")
        self.set_text(
            self.feedback,
            "Avvia un nuovo esame quando vuoi. Le statistiche di estrazione sono già state aggiornate.",
        )
        for child in self.options_frame.winfo_children():
            child.destroy()
        self.submit_button.configure(state="disabled")
        self.next_button.configure(state="disabled")

    def show_stats(self) -> None:
        total_seen = sum(q.get("seen", 0) for q in self.stats.get("questions", {}).values())
        total_correct = sum(q.get("correct", 0) for q in self.stats.get("questions", {}).values())
        total_wrong = sum(q.get("wrong", 0) for q in self.stats.get("questions", {}).values())
        messagebox.showinfo(
            "Statistiche",
            (
                f"Domande nel database: {len(self.questions)}\n"
                f"Estrazioni totali: {total_seen}\n"
                f"Risposte corrette: {total_correct}\n"
                f"Risposte sbagliate: {total_wrong}\n"
                f"Sync: {'online' if self.sync.online else 'offline'}\n"
                f"File statistiche: {STATS_PATH.name}"
            ),
        )

    def reset_counters(self) -> None:
        if not messagebox.askyesno(
            "Reset counter",
            "Vuoi azzerare contatori di estrazione, corrette e sbagliate?",
        ):
            return
        self.stats = {"questions": {}}
        save_stats(self.stats)
        if self.sync.online:
            zero_stats = {
                question.qid: {"seen": 0, "correct": 0, "wrong": 0}
                for question in self.questions
            }
            try:
                self.sync.upsert_stats(zero_stats)
            except SupabaseError as error:
                messagebox.showwarning(
                    "Reset counter",
                    f"Reset locale completato, ma reset Supabase non riuscito:\n{error}",
                )
                self.update_sync_status("offline, reset remoto fallito")
                self.new_session()
                return
        messagebox.showinfo("Reset counter", "Statistiche azzerate.")
        self.new_session()

    @staticmethod
    def set_text(widget: tk.Text, value: str) -> None:
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", value)
        widget.configure(state="disabled")


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"File non trovato: {WORKBOOK_PATH}")
    app = ExamTrainer()
    app.mainloop()


if __name__ == "__main__":
    main()
