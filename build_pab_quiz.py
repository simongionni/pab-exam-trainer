import json
import pathlib
import re
import subprocess
import textwrap
import zlib
import difflib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image


ROOT = pathlib.Path(__file__).resolve().parent
PDFS = [
    ("s1", pathlib.Path(r"C:\Users\simon\Downloads\pab-s1.pdf")),
    ("s2", pathlib.Path(r"C:\Users\simon\Downloads\pab-s2.pdf")),
    ("s3", pathlib.Path(r"C:\Users\simon\Downloads\pab-s3.pdf")),
    ("s4", pathlib.Path(r"C:\Users\simon\Downloads\pab-s4.pdf")),
    ("s5", pathlib.Path(r"C:\Users\simon\Downloads\pab-s5.pdf")),
    ("s6", pathlib.Path(r"C:\Users\simon\Downloads\pab-s6.pdf")),
]

ANSWER_OVERRIDES = {
    ("s4", 214): "C",
    ("s4", 234): "B",
    ("s4", 235): "B",
    ("s4", 236): "C",
    ("s4", "260:"): "CDF",
    ("s4", 260): "A",
    ("s5", 6): "CD",
    ("s5", 27): "AB",
    ("s5", 30): "D",
    ("s5", 37): "C",
    ("s5", 48): "AB",
    ("s5", 54): "A",
    ("s5", 65): "CDE",
}

SKIP_QUESTIONS = {
    ("s5", 1),
    ("s5", 4),
    ("s5", 45),
    ("s5", 63),
}
SKIP_QUESTIONS.update(("s6", number) for number in range(1, 66))


def check_formula(row: int) -> str:
    cleaned = f'SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(D{row},",",""),";","")," ","")'
    return f'=IF(D{row}="","",IF(UPPER({cleaned})=C{row},"OK","NO"))'


def extract_images(pdf_path: pathlib.Path, output_dir: pathlib.Path) -> None:
    output_dir.mkdir(exist_ok=True)
    data = pdf_path.read_bytes()
    page = 1
    pattern = rb"(\d+) 0 obj\s*<<(.*?)>>\s*stream\r?\n(.*?)\r?\nendstream"
    for match in re.finditer(pattern, data, re.S):
        header, stream = match.group(2), match.group(3)
        if b"/Subtype /Image" not in header:
            continue
        width = int(re.search(rb"/Width (\d+)", header).group(1))
        height = int(re.search(rb"/Height (\d+)", header).group(1))
        raw = zlib.decompress(stream)
        mode = "RGB" if len(raw) == width * height * 3 else "L"
        Image.frombytes(mode, (width, height), raw).save(output_dir / f"page_{page:02}.png")
        page += 1


def ocr_pages(image_dir: pathlib.Path, json_path: pathlib.Path) -> None:
    ps_script = rf"""
$ErrorActionPreference='Stop'
Add-Type -AssemblyName System.Runtime.WindowsRuntime
$null = [Windows.Media.Ocr.OcrEngine, Windows.Foundation, ContentType=WindowsRuntime]
$null = [Windows.Graphics.Imaging.BitmapDecoder, Windows.Foundation, ContentType=WindowsRuntime]
$null = [Windows.Storage.StorageFile, Windows.Foundation, ContentType=WindowsRuntime]
$asTaskGeneric = ([System.WindowsRuntimeSystemExtensions].GetMethods() | Where-Object {{ $_.Name -eq 'AsTask' -and $_.GetParameters().Count -eq 1 -and $_.IsGenericMethod }})[0]
function Await($op, $type) {{
  $m=$asTaskGeneric.MakeGenericMethod($type)
  $task=$m.Invoke($null,@($op))
  $task.Wait()
  $task.Result
}}
$engine = [Windows.Media.Ocr.OcrEngine]::TryCreateFromUserProfileLanguages()
$pages = @()
Get-ChildItem -LiteralPath '{image_dir}' -Filter 'page_*.png' | Sort-Object Name | ForEach-Object {{
  $file = Await ([Windows.Storage.StorageFile]::GetFileFromPathAsync($_.FullName)) ([Windows.Storage.StorageFile])
  $stream = Await ($file.OpenReadAsync()) ([Windows.Storage.Streams.IRandomAccessStreamWithContentType])
  $decoder = Await ([Windows.Graphics.Imaging.BitmapDecoder]::CreateAsync($stream)) ([Windows.Graphics.Imaging.BitmapDecoder])
  $bitmap = Await ($decoder.GetSoftwareBitmapAsync()) ([Windows.Graphics.Imaging.SoftwareBitmap])
  $result = Await ($engine.RecognizeAsync($bitmap)) ([Windows.Media.Ocr.OcrResult])
  $lines = @()
  foreach ($line in $result.Lines) {{
    $words = @()
    foreach ($word in $line.Words) {{
      $r = $word.BoundingRect
      $words += [pscustomobject]@{{text=$word.Text; x=$r.X; y=$r.Y; w=$r.Width; h=$r.Height}}
    }}
    $lines += [pscustomobject]@{{text=$line.Text; words=$words}}
  }}
  $pages += [pscustomobject]@{{page=[int]$_.BaseName.Substring(5); image=$_.FullName; lines=$lines}}
}}
$pages | ConvertTo-Json -Depth 8 | Set-Content -LiteralPath '{json_path}' -Encoding UTF8
"""
    subprocess.run(["powershell", "-NoProfile", "-Command", ps_script], check=True)


def normalize_line(text: str) -> str:
    text = re.sub(r"^IO\b", "10", text.strip())
    text = re.sub(r"^([a-e])\.", lambda m: f"{m.group(1).upper()}.", text)
    return text


def clean_s5_line(text: str) -> tuple[str, str | None]:
    text = text.strip()
    selected = None
    match = re.match(r"^Question\s+([IVXLCDM]+)\s+of\s+(\d+)", text, re.I)
    if match:
        roman_text = match.group(1).upper()
        roman = 11 if roman_text == "II" else roman_to_int(roman_text)
        if roman:
            return f"Question {roman} of {match.group(2)}", None

    match = re.match(r"^(@|E|g|O|0|o|C\)|CI|C\])\s+([A-E])\.\s*(.*)$", text)
    if match:
        selected = match.group(2) if match.group(1) in {"@", "E", "g"} else None
        return f"{match.group(2)}. {match.group(3)}".strip(), selected

    return text, None


def roman_to_int(value: str) -> int | None:
    numerals = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    if not value or any(ch not in numerals for ch in value):
        return None
    total = 0
    previous = 0
    for ch in reversed(value):
        current = numerals[ch]
        if current < previous:
            total -= current
        else:
            total += current
            previous = current
    return total


def merged_ocr_lines(json_path: pathlib.Path) -> list[dict]:
    pages = json.loads(json_path.read_text(encoding="utf-8-sig"))
    output: list[dict] = []

    for page in pages:
        output.append({"page": page["page"], "text": f"===== PAGE {page['page']:02} =====", "y": 0, "words": []})
        raw_lines = []
        for line in page["lines"]:
            if not line["words"]:
                continue
            x = min(word["x"] for word in line["words"])
            y = min(word["y"] for word in line["words"])
            raw_lines.append({"text": normalize_line(line["text"]), "x": x, "y": y, "words": line["words"]})

        groups: list[list[dict]] = []
        for line in sorted(raw_lines, key=lambda item: (item["y"], item["x"])):
            if groups and abs(line["y"] - groups[-1][0]["y"]) <= 8:
                groups[-1].append(line)
            else:
                groups.append([line])

        for group in groups:
            ordered = sorted(group, key=lambda item: item["x"])
            text = " ".join(item["text"] for item in ordered if item["text"])
            selected = None
            if json_path.stem.startswith(("pab-s5-", "pab-s6-")):
                text, selected = clean_s5_line(text)
            words = [word for item in ordered for word in item["words"]]
            output.append({
                "page": page["page"],
                "text": text,
                "y": min(item["y"] for item in group),
                "words": words,
                "selected": selected,
            })
        output.append({"page": page["page"], "text": "", "y": 0, "words": []})

    return output


def write_lines(lines: list[dict], path: pathlib.Path) -> None:
    path.write_text("\n".join(line["text"] for line in lines), encoding="utf-8")


def parse_questions(lines: list[dict], source: str) -> list[dict]:
    if source in {"s5", "s6"}:
        return parse_s5_questions(lines, source)

    questions: list[dict] = []
    current = None

    for item in lines:
        line = item["text"].strip()
        if not line or line.startswith("=====") or line in {"Salesforce Certified Platform", "App Builder"}:
            continue

        special_match = re.match(r"^(\d+ of \d+)\.\s+(.+)$", line)
        if special_match:
            number = special_match.group(1)
            rest = special_match.group(2).strip()
            current = {"original_num": number, "source": source, "lines": [rest]}
            questions.append(current)
            continue

        colon_match = re.match(r"^(\d{1,3}):(?:\s+(.+))?$", line)
        if colon_match and valid_question_number(source, int(colon_match.group(1))):
            number = f"{colon_match.group(1)}:"
            rest = (colon_match.group(2) or "").strip()
            current = {"original_num": number, "source": source, "lines": []}
            if rest:
                current["lines"].append(rest)
            questions.append(current)
            continue

        match = re.match(r"^(\d{1,3})\.?(?:\s+(.+))?$", line)
        if match and valid_question_number(source, int(match.group(1))):
            number = int(match.group(1))
            rest = (match.group(2) or "").strip()
            if (
                number == 5
                and questions
                and questions[-1]["lines"]
                and questions[-1]["lines"][-1].startswith(
                    "Sales reps at Universal Containers create multiple quotes"
                )
            ):
                preface = questions[-1]["lines"].pop()
                current = {"original_num": number, "source": source, "lines": [preface]}
            else:
                current = {"original_num": number, "source": source, "lines": []}
            if rest:
                current["lines"].append(rest)
            questions.append(current)
            continue

        if current is not None:
            current["lines"].append(line)

    for question in questions:
        text = "\n".join(question["lines"])
        text = re.sub(r"[ \t]+", " ", text)
        question["text"] = text.strip()
        question["key"] = question_key(question["text"])

    return questions


def parse_s5_questions(lines: list[dict], source: str) -> list[dict]:
    questions: list[dict] = []
    current = None

    for item in lines:
        line = item["text"].strip()
        if not line or line.startswith("====="):
            continue

        match = re.match(r"^Question\s+(\d+)\s+of\s+\d+", line, re.I)
        if match:
            number = int(match.group(1))
            current = {"original_num": number, "source": source, "lines": []}
            questions.append(current)
            continue

        match = re.match(r"^uestlon\s+(\d+)\s+0", line, re.I)
        if match:
            number = int(match.group(1))
            current = {"original_num": number, "source": source, "lines": []}
            questions.append(current)
            continue

        compact = re.sub(r"[^a-z]", "", line.lower())
        if compact in {"uesiono", "ueslono", "questiono"} and current is not None:
            number = int(current["original_num"]) + 1
            current = {"original_num": number, "source": source, "lines": []}
            questions.append(current)
            continue

        if current is None:
            continue

        if re.match(r"^Choose\s+\d+\s+option", line, re.I):
            continue

        current["lines"].append(line)

    for question in questions:
        text = "\n".join(question["lines"])
        text = re.sub(r"[ \t]+", " ", text)
        question["text"] = text.strip()
        question["key"] = question_key(question["text"])

    return questions


def valid_question_number(source: str, number: int) -> bool:
    if source == "s1":
        return 1 <= number <= 65
    if source == "s2":
        return 66 <= number <= 131
    if source == "s3":
        return 132 <= number <= 200
    if source == "s4":
        return 201 <= number <= 999
    if source in {"s5", "s6"}:
        return 1 <= number <= 65
    return 1 <= number <= 999


def question_key(text: str) -> str:
    stem = re.split(r"\nA\.", text, maxsplit=1)[0]
    stem = re.sub(r"[^a-z0-9]+", " ", stem.lower()).strip()
    return stem


def highlighted_options(lines: list[dict], image_dir: pathlib.Path, source: str) -> dict[object, str]:
    if source in {"s5", "s6"}:
        return selected_radio_options(lines, image_dir)

    masks = {}
    for image_path in sorted(image_dir.glob("page_*.png")):
        page = int(image_path.stem.split("_")[1])
        image = Image.open(image_path).convert("RGB")
        pix = image.load()
        masks[page] = (image.size, pix)

    answers: dict[object, set[str]] = {}
    current_q = None
    current_option = None

    for item in lines:
        text = item["text"].strip()
        if not text or text.startswith("=====") or text in {"Salesforce Certified Platform", "App Builder"}:
            continue

        colon_match = re.match(r"^(\d{1,3}):", text)
        if colon_match and valid_question_number(source, int(colon_match.group(1))):
            current_q = f"{colon_match.group(1)}:"
            current_option = None

        special_match = re.match(r"^(\d+ of \d+)\.", text)
        if special_match:
            current_q = special_match.group(1)
            current_option = None

        q_match = re.match(r"^(\d{1,3})\b", text)
        if not colon_match and q_match and valid_question_number(source, int(q_match.group(1))):
            current_q = int(q_match.group(1))
            current_option = None

        opt_match = re.match(r"^([A-F])\.", text)
        if opt_match:
            current_option = opt_match.group(1)

        if current_q and current_option and is_highlighted(item, masks):
            answers.setdefault(current_q, set()).add(current_option)

    return {number: "".join(sorted(options)) for number, options in answers.items()}


def is_highlighted(item: dict, masks: dict) -> bool:
    if not item["words"] or item["page"] not in masks:
        return False

    (width, height), pix = masks[item["page"]]
    yellow = 0
    total = 0
    for word in item["words"]:
        x0 = max(0, int(word["x"]) - 2)
        y0 = max(0, int(word["y"]) - 2)
        x1 = min(width, int(word["x"] + word["w"]) + 2)
        y1 = min(height, int(word["y"] + word["h"]) + 2)
        step_x = max(1, (x1 - x0) // 10)
        step_y = max(1, (y1 - y0) // 4)
        for y in range(y0, y1, step_y):
            for x in range(x0, x1, step_x):
                r, g, b = pix[x, y]
                total += 1
                if r > 220 and g > 180 and b < 90:
                    yellow += 1

    return total > 0 and yellow / total > 0.03


def selected_radio_options(lines: list[dict], image_dir: pathlib.Path) -> dict[object, str]:
    answers: dict[object, str] = {}
    current_q = None
    radio_masks = masks_for_radio(image_dir)

    for item in lines:
        text = item["text"].strip()
        if not text or text.startswith("====="):
            continue

        q_match = re.match(r"^Question\s+(\d+)\s+of\s+\d+", text, re.I)
        if q_match:
            current_q = int(q_match.group(1))
            continue

        if current_q and item.get("selected"):
            answers[current_q] = item["selected"]
            continue

        opt_match = re.match(r"^([A-F])(?:\.|\s)\s*", text)
        if current_q and opt_match and has_blue_radio_near_option(item, radio_masks):
            answers[current_q] = "".join(sorted(set(answers.get(current_q, "") + opt_match.group(1))))

    return answers


def masks_for_radio(image_dir: pathlib.Path) -> dict:
    masks = {}
    for image_path in sorted(image_dir.glob("page_*.png")):
        page = int(image_path.stem.split("_")[1])
        image = Image.open(image_path).convert("RGB")
        masks[page] = (image.size, image.load())
    return masks


def has_blue_radio_near_option(item: dict, masks: dict) -> bool:
    if not item["words"] or item["page"] not in masks:
        return False

    (width, height), pix = masks[item["page"]]
    min_x = min(word["x"] for word in item["words"])
    min_y = min(word["y"] for word in item["words"])
    max_y = max(word["y"] + word["h"] for word in item["words"])

    x0 = max(0, int(min_x) - 55)
    x1 = max(0, int(min_x) - 5)
    y0 = max(0, int(min_y) - 8)
    y1 = min(height, int(max_y) + 8)

    blue = 0
    total = 0
    for y in range(y0, y1):
        for x in range(x0, min(width, x1)):
            r, g, b = pix[x, y]
            total += 1
            if b > 150 and g > 90 and r < 120:
                blue += 1

    return total > 0 and blue / total > 0.01


def process_pdf(source: str, pdf_path: pathlib.Path) -> tuple[list[dict], dict[int, str], pathlib.Path]:
    image_dir = ROOT / f"pab_{source}_pages"
    json_path = ROOT / f"pab-{source}-ocr-words.json"
    lines_path = ROOT / f"pab-{source}-ocr-lines.txt"

    extract_images(pdf_path, image_dir)
    ocr_pages(image_dir, json_path)
    lines = merged_ocr_lines(json_path)
    write_lines(lines, lines_path)

    return parse_questions(lines, source), highlighted_options(lines, image_dir, source), lines_path


def build_workbook(questions: list[dict], output: pathlib.Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quiz"
    sheet.append(["N", "Fonte", "N originale", "Domanda e opzioni (OCR)", "Corretta", "La tua risposta", "Esito", "Note"])

    for number, question in enumerate(questions, start=1):
        row = sheet.max_row + 1
        note = question.get("note", "")
        sheet.append([
            number,
            question["source"],
            question["original_num"],
            question["text"],
            question.get("answer", ""),
            "",
            check_formula_for_columns(row),
            note,
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = {"A": 6, "B": 9, "C": 11, "D": 100, "E": 12, "F": 16, "G": 10, "H": 40}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for idx in [3, 7]:
            row[idx].alignment = Alignment(wrap_text=True, vertical="top")
        for idx in [0, 1, 2, 4, 5, 6]:
            row[idx].alignment = Alignment(horizontal="center", vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:H{sheet.max_row}"

    instructions = workbook.create_sheet("Istruzioni")
    instructions["A1"] = "Come usarlo"
    instructions["A1"].font = Font(bold=True, size=14)
    instructions["A2"] = (
        "Inserisci la tua risposta nella colonna F del foglio Quiz. "
        "Per domande a risposta multipla usa lettere in ordine alfabetico, es. AD o BCE. "
        "La colonna G mostra OK/NO."
    )
    instructions["A3"] = (
        "Le risposte corrette sono ricavate dalle evidenziazioni gialle nei PDF. "
        "La trascrizione deriva da OCR, quindi controlla nel PDF le righe con formule degradate."
    )
    instructions.column_dimensions["A"].width = 120
    workbook.save(output)


def check_formula_for_columns(row: int) -> str:
    cleaned = f'SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(F{row},",",""),";","")," ","")'
    return f'=IF(F{row}="","",IF(UPPER({cleaned})=E{row},"OK","NO"))'


def main() -> None:
    all_questions: list[dict] = []
    seen = set()
    seen_keys: list[str] = []
    added_by_source = {}
    paths = []

    for source, pdf_path in PDFS:
        questions, answers, lines_path = process_pdf(source, pdf_path)
        paths.append(lines_path)
        added = 0
        for question in questions:
            if (source, question["original_num"]) in SKIP_QUESTIONS:
                continue
            key = question["key"]
            if key in seen or is_near_duplicate(key, seen_keys):
                continue
            seen.add(key)
            seen_keys.append(key)
            question["answer"] = ANSWER_OVERRIDES.get(
                (source, question["original_num"]),
                answers.get(question["original_num"], ""),
            )
            if not question["answer"]:
                question["note"] = "Risposta evidenziata non rilevata automaticamente."
            all_questions.append(question)
            added += 1
        added_by_source[source] = added

    combined_lines = []
    for source, _pdf_path in PDFS:
        path = ROOT / f"pab-{source}-ocr-lines.txt"
        combined_lines.append(f"===== {source.upper()} =====")
        combined_lines.append(path.read_text(encoding="utf-8").strip())
        combined_lines.append("")
    (ROOT / "pab-combined-ocr-lines.txt").write_text("\n".join(combined_lines), encoding="utf-8")

    final_lines = []
    for number, question in enumerate(all_questions, start=1):
        final_lines.append(
            f"{number} [fonte: {question['source']}, originale: {question['original_num']}] "
            f"{question['text']}"
        )
        final_lines.append("")
    (ROOT / "pab-s1-ocr-lines.txt").write_text("\n".join(final_lines).rstrip() + "\n", encoding="utf-8")

    build_workbook(all_questions, ROOT / "pab-s1-quiz.xlsx")
    print(f"saved pab-s1-quiz.xlsx with {len(all_questions)} unique questions")
    print(json.dumps(added_by_source, ensure_ascii=False))


def is_near_duplicate(key: str, seen_keys: list[str]) -> bool:
    return any(difflib.SequenceMatcher(None, key, existing).ratio() >= 0.96 for existing in seen_keys)


if __name__ == "__main__":
    main()
