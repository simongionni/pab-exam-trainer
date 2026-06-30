import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "pab-s1-quiz.xlsx"
OUTPUT_PATH = ROOT / "supabase" / "seed_questions.sql"


def normalize_answer(value: str) -> str:
    return "".join(sorted(set(re.sub(r"[^A-F]", "", str(value or "").upper()))))


def split_question_text(text: str) -> tuple[str, list[dict[str, str]]]:
    lines = [line.rstrip() for line in str(text or "").splitlines()]
    prompt_lines: list[str] = []
    options: list[tuple[str, list[str]]] = []
    current = None

    option_re = re.compile(r"^\s*([A-F])[\.\)]\s*(.*)$")
    for line in lines:
        match = option_re.match(line)
        if match:
            current = [match.group(1), [match.group(2).strip()]]
            options.append(current)
            continue
        if current is not None:
            current[1].append(line.strip())
        else:
            prompt_lines.append(line.strip())

    prompt = "\n".join(line for line in prompt_lines if line).strip()
    clean_options = []
    for letter, parts in options:
        clean_options.append(
            {
                "letter": letter,
                "text": " ".join(part for part in parts if part).strip(),
            }
        )
    return prompt, clean_options


def sql_string(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def sql_jsonb(value: object) -> str:
    return sql_string(json.dumps(value, ensure_ascii=False)) + "::jsonb"


def sql_text_array(values: list[str]) -> str:
    return "array[" + ", ".join(sql_string(value) for value in values) + "]::text[]"


def build_rows() -> list[str]:
    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    sheet = workbook["Quiz"]
    rows = []

    for row in range(2, sheet.max_row + 1):
        qid = str(sheet.cell(row, 1).value or "").strip()
        source = str(sheet.cell(row, 2).value or "").strip()
        original_number = str(sheet.cell(row, 3).value or "").strip()
        text = sheet.cell(row, 4).value
        correct = list(normalize_answer(sheet.cell(row, 5).value))
        prompt, options = split_question_text(text)

        if not qid or not prompt or not options or not correct:
            continue

        rows.append(
            "("
            + ", ".join(
                [
                    sql_string(qid),
                    sql_string(source),
                    sql_string(original_number),
                    sql_string(prompt),
                    sql_jsonb(options),
                    sql_text_array(correct),
                ]
            )
            + ")"
        )

    return rows


def main() -> None:
    rows = build_rows()
    sql = """-- Generated from pab-s1-quiz.xlsx. Safe to run multiple times.
begin;

insert into public.questions (
  id,
  source,
  original_number,
  prompt,
  options,
  correct_letters
)
values
"""
    sql += ",\n".join(rows)
    sql += """
on conflict (id) do update set
  source = excluded.source,
  original_number = excluded.original_number,
  prompt = excluded.prompt,
  options = excluded.options,
  correct_letters = excluded.correct_letters;

commit;
"""

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(sql, encoding="utf-8")
    print(f"Generated {len(rows)} question rows in {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
